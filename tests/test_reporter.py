import sys
import tempfile
import unittest
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, PatternFill, Side

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / 'src'))

from reporter import ExcelReporter


class ExcelReporterTests(unittest.TestCase):
    def _create_template(self, directory: str) -> Path:
        path = Path(directory) / 'template.xlsx'
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = '様式'
        sheet['F1'] = '氏　名'
        sheet.merge_cells('G1:H1')
        headers = ['No.', '実施年月日', 'ＳＮＳ別', 'サイト名・ユーザー名', 'Ｕ　Ｒ　Ｌ', '該当項目', '備考']
        for column, value in enumerate(headers, start=2):
            sheet.cell(4, column, value)
        thin = Side(style='thin', color='000000')
        for row in range(7, 9):
            sheet.cell(row, 2, row - 6)
            for column in range(2, 9):
                cell = sheet.cell(row, column)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.fill = PatternFill('solid', fgColor='FFF2CC')
                cell.alignment = Alignment(vertical='center', wrap_text=True)
            sheet.row_dimensions[row].height = 39
        workbook.save(path)
        return path

    def test_writes_to_first_empty_numbered_row_without_losing_style(self):
        with tempfile.TemporaryDirectory() as directory:
            path = self._create_template(directory)
            reporter = ExcelReporter(str(path), reporter_name='山田 太郎')
            reporter.append_record(
                url='https://dangerous.example/login',
                target_brand='PayPay',
                features='ロゴと認証画面を悪用',
                detected_at='2026-09-01T12:00:00+09:00',
                category='フィッシングサイト',
            )

            self.assertEqual(reporter._next_row, 8)
            self.assertEqual(reporter._sheet['G1'].value, '山田 太郎')
            self.assertEqual(reporter._sheet['F7'].value, 'https://dangerous.example/login')
            self.assertIsNone(reporter._sheet['F7'].hyperlink)
            self.assertEqual(reporter._sheet['F7'].fill.fgColor.rgb, '00FFF2CC')
            self.assertTrue(reporter._sheet['F7'].alignment.wrap_text)
            self.assertEqual(reporter._sheet.row_dimensions[7].height, 39)

    def test_legacy_float_sequence_is_recognized(self):
        self.assertTrue(ExcelReporter._is_sequence_number(1.0))
        self.assertFalse(ExcelReporter._is_sequence_number(1.5))
        self.assertFalse(ExcelReporter._is_sequence_number(True))

    def test_full_width_spaced_header_is_normalized(self):
        self.assertEqual(ExcelReporter._normalize_label('Ｕ　Ｒ　Ｌ'), 'url')

    def test_missing_template_fails_instead_of_creating_wrong_workbook(self):
        with tempfile.TemporaryDirectory() as directory:
            missing = Path(directory) / 'missing.xlsx'
            with self.assertRaises(FileNotFoundError):
                ExcelReporter(str(missing))

    def test_report_is_saved_in_separate_output_directory(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            template_dir = root / 'templates'
            output_dir = root / 'detections'
            template_dir.mkdir()
            template = self._create_template(str(template_dir))
            reporter = ExcelReporter(
                str(template), reporter_name='山田 太郎', output_dir=str(output_dir)
            )
            reporter.append_record(
                url='https://dangerous.example', target_brand='PayPay',
                features='確認済み証拠', category='フィッシングサイト',
            )

            saved = Path(reporter.save())

            self.assertEqual(saved.parent, output_dir)
            self.assertTrue(saved.exists())
            self.assertTrue(template.exists())
            self.assertEqual(len(list(template_dir.glob('template_*.xlsx'))), 0)

    def test_report_can_be_saved_directly_to_selected_path(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            template = self._create_template(directory)
            selected = root / 'chosen' / '確認結果.xlsx'
            reporter = ExcelReporter(str(template), reporter_name='確認者')
            reporter.append_record(
                url='https://dangerous.example', target_brand='確認対象',
                features='人手確認済み', category='phishing',
            )

            saved = Path(reporter.save(selected))

            self.assertEqual(saved, selected)
            self.assertTrue(saved.exists())
            self.assertGreater(saved.stat().st_size, 0)

    def test_template_cannot_be_overwritten(self):
        with tempfile.TemporaryDirectory() as directory:
            template = self._create_template(directory)
            reporter = ExcelReporter(str(template))
            with self.assertRaisesRegex(ValueError, 'テンプレート'):
                reporter.save(template)


if __name__ == '__main__':
    unittest.main()
