"""Thread-safe, append-only audit logs for candidate and scan URLs."""

from __future__ import annotations

import datetime as dt
import csv
import json
import logging
import threading
import uuid
from pathlib import Path
from typing import Any

from url_normalization import safe_observation_url


logger = logging.getLogger(__name__)
FILTER_PASSED_LOG_NAME = 'filter_passed_urls.log'
SCANNED_LOG_NAME = 'scanned_urls.log'
FILTER_PASSED_CSV_NAME = 'filter_passed_urls.csv'
SCANNED_CSV_NAME = 'scanned_urls.csv'
CSV_HEADERS = (
    '日時', '処理', 'URL', 'ドメイン', '収集元', '候補分類', 'スコア', 'セッションID'
)
EVENT_LABELS = {
    'session_started': '監視開始',
    'filter_passed': 'フィルタ通過',
    'scan_started': 'スキャン開始',
}
CANDIDATE_KIND_LABELS = {
    'known_phishing': '既知のフィッシング候補',
    'brand_impersonation': 'ブランド偽装候補',
    'counterfeit_goods': 'コピー商品販売候補',
    'suspected_illegal_goods': '違法商品販売候補',
    'suspicious_shop': '不審な通販サイト候補',
}


class UrlAuditLog:
    """Write two human-readable JSON Lines files without URL secrets."""

    def __init__(self, log_dir: str | Path | None = None):
        self.log_dir = (
            Path(log_dir)
            if log_dir is not None
            else Path(__file__).resolve().parents[1] / 'logs'
        )
        self.filter_passed_path = self.log_dir / FILTER_PASSED_LOG_NAME
        self.scanned_path = self.log_dir / SCANNED_LOG_NAME
        self.filter_passed_csv_path = self.log_dir / FILTER_PASSED_CSV_NAME
        self.scanned_csv_path = self.log_dir / SCANNED_CSV_NAME
        self.session_id = str(uuid.uuid4())
        self._lock = threading.Lock()
        self._enabled = True
        self._failure_reported = False
        self._start_session()

    @property
    def paths(self) -> tuple[Path, Path]:
        return self.filter_passed_path, self.scanned_path

    @property
    def human_paths(self) -> tuple[Path, Path]:
        """CSV logs intended for people and direct opening in Excel."""
        return self.filter_passed_csv_path, self.scanned_csv_path

    def _start_session(self) -> None:
        try:
            self.log_dir.mkdir(parents=True, exist_ok=True)
            for path, csv_path, stream in (
                (self.filter_passed_path, self.filter_passed_csv_path, 'filter_passed'),
                (self.scanned_path, self.scanned_csv_path, 'scanned'),
            ):
                try:
                    self._backfill_human_csv(path, csv_path)
                except OSError as exc:
                    self._report_human_failure(exc)
                self._append(path, {
                    'timestamp': self._timestamp(),
                    'event': 'session_started',
                    'stream': stream,
                    'session_id': self.session_id,
                })
        except OSError as exc:
            self._enabled = False
            self._report_failure(exc)

    def _backfill_human_csv(self, log_path: Path, csv_path: Path) -> None:
        """Build the readable CSV from an older JSON log on first use."""
        if (csv_path.exists() and csv_path.stat().st_size > 0) or not log_path.exists():
            return
        with csv_path.open('w', encoding='utf-8-sig', newline='') as output:
            writer = csv.writer(output)
            writer.writerow(CSV_HEADERS)
            with log_path.open('r', encoding='utf-8') as source:
                for line in source:
                    try:
                        entry = json.loads(line)
                    except (json.JSONDecodeError, TypeError):
                        continue
                    if isinstance(entry, dict):
                        writer.writerow(self._human_row(entry))

    def record_filter_passed(
        self,
        url: str,
        *,
        domain: str,
        source: str,
        candidate_kind: str,
        score: int,
    ) -> bool:
        try:
            safe_score = max(0, min(int(score), 100))
        except (TypeError, ValueError):
            safe_score = 0
        return self._record(self.filter_passed_path, {
            'event': 'filter_passed',
            'url': url,
            'domain': str(domain)[:253],
            'source': str(source)[:64],
            'candidate_kind': str(candidate_kind)[:64],
            'score': safe_score,
        })

    def record_scanned(
        self,
        url: str,
        *,
        domain: str,
        source: str,
        candidate_kind: str,
    ) -> bool:
        return self._record(self.scanned_path, {
            'event': 'scan_started',
            'url': url,
            'domain': str(domain)[:253],
            'source': str(source)[:64],
            'candidate_kind': str(candidate_kind)[:64],
        })

    def _record(self, path: Path, values: dict[str, Any]) -> bool:
        if not self._enabled:
            return False
        try:
            safe_url = safe_observation_url(values['url'])
            entry = {
                'timestamp': self._timestamp(),
                'session_id': self.session_id,
                **values,
                'url': safe_url,
            }
            self._append(path, entry)
            return True
        except (OSError, TypeError, ValueError) as exc:
            self._report_failure(exc)
            return False

    def _append(self, path: Path, entry: dict[str, Any]) -> None:
        line = json.dumps(entry, ensure_ascii=False, separators=(',', ':'))
        csv_path = (
            self.filter_passed_csv_path
            if path == self.filter_passed_path
            else self.scanned_csv_path
        )
        with self._lock:
            with path.open('a', encoding='utf-8', newline='\n') as handle:
                handle.write(line + '\n')
            try:
                write_header = not csv_path.exists() or csv_path.stat().st_size == 0
                with csv_path.open('a', encoding='utf-8-sig', newline='') as handle:
                    writer = csv.writer(handle)
                    if write_header:
                        writer.writerow(CSV_HEADERS)
                    writer.writerow(self._human_row(entry))
            except OSError as exc:
                self._report_human_failure(exc)

    @staticmethod
    def _human_row(entry: dict[str, Any]) -> list[str | int]:
        timestamp = str(entry.get('timestamp', '')).replace('T', ' ', 1)
        candidate_kind = str(entry.get('candidate_kind', ''))
        values = [
            timestamp,
            EVENT_LABELS.get(str(entry.get('event', '')), entry.get('event', '')),
            entry.get('url', ''),
            entry.get('domain', ''),
            entry.get('source', ''),
            CANDIDATE_KIND_LABELS.get(candidate_kind, candidate_kind),
            entry.get('score', ''),
            entry.get('session_id', ''),
        ]
        return [UrlAuditLog._safe_csv_value(value) for value in values]

    @staticmethod
    def _safe_csv_value(value: Any) -> str:
        text = str(value)
        if text.lstrip().startswith(('=', '+', '-', '@')):
            return "'" + text
        return text

    def _report_failure(self, exc: Exception) -> None:
        if not self._failure_reported:
            logger.warning('URL監査ログを書き込めません。監視処理は継続します: %s', exc)
            self._failure_reported = True

    def _report_human_failure(self, exc: Exception) -> None:
        if not self._failure_reported:
            logger.warning(
                '人向けCSVログを書き込めません。CSVをExcelで開いている場合は閉じてください。'
                '監査用JSONログは継続します: %s',
                exc,
            )
            self._failure_reported = True

    @staticmethod
    def _timestamp() -> str:
        return dt.datetime.now().astimezone().isoformat(timespec='seconds')


def export_audit_logs_to_excel(
    destination: str | Path,
    log_dir: str | Path | None = None,
) -> Path:
    """Create a readable Excel workbook from both append-only JSON logs."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    directory = (
        Path(log_dir)
        if log_dir is not None
        else Path(__file__).resolve().parents[1] / 'logs'
    )
    output_path = Path(destination).expanduser()
    if output_path.suffix.lower() != '.xlsx':
        output_path = output_path.with_suffix('.xlsx')
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)
    sources = (
        ('フィルタ通過URL', directory / FILTER_PASSED_LOG_NAME),
        ('スキャン対象URL', directory / SCANNED_LOG_NAME),
    )
    for sheet_name, source_path in sources:
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(CSV_HEADERS)
        if source_path.exists():
            with source_path.open('r', encoding='utf-8') as handle:
                for line in handle:
                    try:
                        entry = json.loads(line)
                    except (json.JSONDecodeError, TypeError):
                        continue
                    if isinstance(entry, dict):
                        sheet.append(UrlAuditLog._human_row(entry))
        for cell in sheet[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='1F4E78')
            cell.alignment = Alignment(horizontal='center')
        sheet.freeze_panes = 'A2'
        sheet.auto_filter.ref = sheet.dimensions
        widths = (25, 16, 70, 32, 16, 25, 10, 38)
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + index)].width = width
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
            row[2].hyperlink = None

    workbook.save(output_path)
    return output_path
