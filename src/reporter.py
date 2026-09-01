import logging
import os
from copy import copy
from datetime import datetime
from numbers import Real
from pathlib import Path
from typing import Optional
import unicodedata
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
logger = logging.getLogger(__name__)
TARGET_SHEET_NAME = '様式'
COLUMN_KEYWORDS = {'実施年月日': ['実施年月日', '年月日', 'date'], 'ＳＮＳ別': ['ＳＮＳ別', 'sns別', 'sns', '種別'], 'サイト名・ユーザー名': ['サイト名・ユーザー名', 'サイト名', 'ユーザー名', '名称'], 'ＵＲＬ': ['ＵＲＬ', 'url', 'アドレス'], '該当項目': ['該当項目', '分類', 'カテゴリ'], '備考': ['備考', '特徴', 'メモ']}
PROJECT_ROOT = Path(__file__).resolve().parent.parent

class ExcelReporter:

    def __init__(self, excel_path: str, reporter_name: str = ''):
        self._excel_path = self._resolve_excel_path(excel_path)
        self._reporter_name = reporter_name.strip()
        self._workbook: Optional[openpyxl.Workbook] = None
        self._sheet = None
        self._column_map: dict[str, int] = {}
        self._first_data_row: int = 0
        self._next_row: int = 0
        self._appended_count: int = 0
        # Prefer the faithfully converted .xlsx template when both formats exist.
        if self._excel_path.suffix.lower() == '.xls':
            xlsx_path = self._excel_path.with_suffix('.xlsx')
            if xlsx_path.exists():
                logger.info(f'.xlsx テンプレートを優先します: {xlsx_path}')
                self._excel_path = xlsx_path
        if self._excel_path.suffix.lower() == '.xls':
            self._output_path = self._excel_path.with_suffix('.xlsx')
        else:
            self._output_path = self._excel_path
        self._load_workbook()

    @staticmethod
    def _resolve_excel_path(excel_path: str) -> Path:
        """Resolve GUI/CLI relative paths independently of the current directory."""
        path = Path(excel_path).expanduser()
        if path.is_absolute():
            return path
        for candidate in (PROJECT_ROOT / path, Path.cwd() / path):
            if candidate.exists():
                return candidate.resolve()
        return (PROJECT_ROOT / path).resolve()

    def _load_workbook(self) -> None:
        if not self._excel_path.exists():
            raise FileNotFoundError(f'Excelテンプレートが見つかりません: {self._excel_path}')
        file_ext = self._excel_path.suffix.lower()
        try:
            if file_ext == '.xls':
                self._load_from_xls()
            elif file_ext in ('.xlsx', '.xlsm'):
                self._workbook = openpyxl.load_workbook(
                    str(self._excel_path), keep_vba=file_ext == '.xlsm'
                )
                self._setup_sheet()
            else:
                raise ValueError(f'サポートされていないExcel形式です: {file_ext}')
        except Exception as e:
            logger.error(f'Excel 読み込みエラー: {e}')
            raise

    def _load_from_xls(self) -> None:
        try:
            import xlrd
        except ImportError:
            logger.error('xlrd がインストールされていません。`pip install xlrd==1.2.0` を実行してください。')
            self._create_new_workbook()
            return
        xls_book = xlrd.open_workbook(str(self._excel_path), formatting_info=True)
        sheet_names = xls_book.sheet_names()
        if TARGET_SHEET_NAME in sheet_names:
            xls_sheet = xls_book.sheet_by_name(TARGET_SHEET_NAME)
        elif sheet_names:
            xls_sheet = xls_book.sheet_by_index(0)
            logger.warning(f"シート '{TARGET_SHEET_NAME}' が見つかりません。最初のシート '{xls_sheet.name}' を使用します。")
        else:
            self._create_new_workbook()
            return
        self._workbook = openpyxl.Workbook()
        ws = self._workbook.active
        ws.title = xls_sheet.name
        self._copy_xls_sheet(xls_book, xls_sheet, ws)
        logger.info('✅ .xls テンプレートの値と書式を読み込みました')
        self._setup_sheet()

    @staticmethod
    def _xls_color(book, color_index: int) -> Optional[str]:
        rgb = book.colour_map.get(color_index)
        return 'FF' + ''.join(f'{component:02X}' for component in rgb) if rgb else None

    @classmethod
    def _copy_xls_sheet(cls, book, source, target) -> None:
        """Copy the legacy form without discarding its layout and cell styles."""
        import xlrd

        border_styles = {
            0: None, 1: 'thin', 2: 'medium', 3: 'dashed', 4: 'dotted',
            5: 'thick', 6: 'double', 7: 'hair', 8: 'mediumDashed',
            9: 'dashDot', 10: 'mediumDashDot', 11: 'dashDotDot',
            12: 'mediumDashDotDot', 13: 'slantDashDot',
        }
        fill_patterns = {
            1: 'solid', 2: 'mediumGray', 3: 'darkGray', 4: 'lightGray',
            5: 'darkHorizontal', 6: 'darkVertical', 7: 'darkDown',
            8: 'darkUp', 9: 'darkGrid', 10: 'darkTrellis',
            11: 'lightHorizontal', 12: 'lightVertical', 13: 'lightDown',
            14: 'lightUp', 15: 'lightGrid', 16: 'lightTrellis',
            17: 'gray125', 18: 'gray0625',
        }
        horizontal = {0: 'general', 1: 'left', 2: 'center', 3: 'right', 4: 'fill', 5: 'justify', 6: 'centerContinuous', 7: 'distributed'}
        vertical = {0: 'top', 1: 'center', 2: 'bottom', 3: 'justify', 4: 'distributed'}

        def side(line_style: int, color_index: int) -> Side:
            return Side(
                style=border_styles.get(line_style),
                color=cls._xls_color(book, color_index),
            )

        for row_idx in range(source.nrows):
            for col_idx in range(source.ncols):
                source_cell = source.cell(row_idx, col_idx)
                value = source_cell.value
                if source_cell.ctype == xlrd.XL_CELL_DATE:
                    converted = xlrd.xldate_as_datetime(value, book.datemode)
                    value = converted.date() if converted.time() == datetime.min.time() else converted
                elif source_cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                    value = None
                cell = target.cell(row=row_idx + 1, column=col_idx + 1, value=value)
                xf = book.xf_list[source_cell.xf_index]
                old_font = book.font_list[xf.font_index]
                cell.font = Font(
                    name=old_font.name,
                    size=old_font.height / 20 if old_font.height else None,
                    bold=bool(old_font.bold), italic=bool(old_font.italic),
                    underline='single' if old_font.underline_type else None,
                    strike=bool(old_font.struck_out),
                    color=cls._xls_color(book, old_font.colour_index),
                )
                pattern = fill_patterns.get(xf.background.fill_pattern)
                if pattern:
                    foreground = cls._xls_color(book, xf.background.pattern_colour_index)
                    background = cls._xls_color(book, xf.background.background_colour_index)
                    cell.fill = PatternFill(
                        patternType=pattern,
                        fgColor=foreground or '00000000',
                        bgColor=background or '00000000',
                    )
                old_border = xf.border
                cell.border = Border(
                    left=side(old_border.left_line_style, old_border.left_colour_index),
                    right=side(old_border.right_line_style, old_border.right_colour_index),
                    top=side(old_border.top_line_style, old_border.top_colour_index),
                    bottom=side(old_border.bottom_line_style, old_border.bottom_colour_index),
                )
                old_alignment = xf.alignment
                cell.alignment = Alignment(
                    horizontal=horizontal.get(old_alignment.hor_align),
                    vertical=vertical.get(old_alignment.vert_align),
                    wrap_text=bool(old_alignment.text_wrapped),
                    text_rotation=old_alignment.rotation,
                    shrink_to_fit=bool(old_alignment.shrink_to_fit),
                    indent=old_alignment.indent_level,
                )
                if xf.format_key in book.format_map:
                    cell.number_format = book.format_map[xf.format_key].format_str

        for row_idx, row_info in source.rowinfo_map.items():
            target.row_dimensions[row_idx + 1].height = row_info.height / 20
            target.row_dimensions[row_idx + 1].hidden = bool(row_info.hidden)
        for col_idx, col_info in source.colinfo_map.items():
            dimension = target.column_dimensions[get_column_letter(col_idx + 1)]
            dimension.width = col_info.width / 256
            dimension.hidden = bool(col_info.hidden)
        for row_start, row_end, col_start, col_end in source.merged_cells:
            target.merge_cells(
                start_row=row_start + 1, end_row=row_end,
                start_column=col_start + 1, end_column=col_end,
            )
        target.sheet_view.showGridLines = bool(source.show_grid_lines)

    def _create_new_workbook(self) -> None:
        self._workbook = openpyxl.Workbook()
        ws = self._workbook.active
        ws.title = TARGET_SHEET_NAME
        headers = ['実施年月日', 'ＳＮＳ別', 'サイト名・ユーザー名', 'ＵＲＬ', '該当項目', '備考']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
        self._output_path = Path('CYCOT_report.xlsx')
        logger.info('新規 Excel ワークブックを作成しました')
        self._setup_sheet()

    def _setup_sheet(self) -> None:
        ws = self._workbook[TARGET_SHEET_NAME] if TARGET_SHEET_NAME in self._workbook.sheetnames else self._workbook.active
        self._sheet = ws
        self._column_map = self._detect_columns()
        self._set_reporter_name()
        self._next_row = self._find_next_data_row()
        logger.info(f"📊 シート '{ws.title}' を使用。データ追記開始行: {self._next_row}")

    def _detect_columns(self) -> dict[str, int]:
        column_map = {}
        ws = self._sheet
        for row in ws.iter_rows(min_row=1, max_row=5):
            for cell in row:
                if cell.value is None:
                    continue
                cell_str = self._normalize_label(cell.value)
                for col_key, keywords in COLUMN_KEYWORDS.items():
                    if col_key in column_map:
                        continue
                    if any(self._normalize_label(keyword) in cell_str for keyword in keywords):
                        column_map[col_key] = cell.column
                        logger.debug(f"列マッピング: '{col_key}' → 列 {cell.column}")
                        break
        numbered_form = ws.max_column >= 8 and any(
            self._is_sequence_number(ws.cell(row=row, column=2).value)
            for row in range(1, ws.max_row + 1)
        )
        if numbered_form and len(column_map) < len(COLUMN_KEYWORDS):
            column_map = {
                '実施年月日': 3,
                'ＳＮＳ別': 4,
                'サイト名・ユーザー名': 5,
                'ＵＲＬ': 6,
                '該当項目': 7,
                '備考': 8,
            }
        elif not column_map:
            logger.warning('列ヘッダーを自動検出できませんでした。デフォルト順を使用します。')
            column_map = {'実施年月日': 1, 'ＳＮＳ別': 2, 'サイト名・ユーザー名': 3, 'ＵＲＬ': 4, '該当項目': 5, '備考': 6}
        return column_map

    def _set_reporter_name(self) -> None:
        if not self._reporter_name:
            return
        name_cell = None
        for row in self._sheet.iter_rows(min_row=1, max_row=min(3, self._sheet.max_row)):
            for cell in row:
                label = str(cell.value or '').replace(' ', '').replace('　', '')
                if label in ('氏名', '名前'):
                    name_cell = self._sheet.cell(row=cell.row, column=cell.column + 1)
                    break
            if name_cell:
                break
        # The supplied CYCOT form places the name value in G1.
        if name_cell is None and self._sheet.max_column >= 7:
            name_cell = self._sheet['G1']
        if name_cell is not None:
            name_cell.value = self._reporter_name

    def _find_next_data_row(self) -> int:
        """Return the first empty numbered row in the supplied form."""
        data_columns = list(self._column_map.values())
        for row in range(1, self._sheet.max_row + 1):
            sequence = self._sheet.cell(row=row, column=2).value
            if self._is_sequence_number(sequence) and all(
                self._sheet.cell(row=row, column=column).value in (None, '')
                for column in data_columns
            ):
                return row
        return self._find_last_data_row() + 1

    @staticmethod
    def _normalize_label(value) -> str:
        normalized = unicodedata.normalize('NFKC', str(value)).lower()
        return ''.join(normalized.split())

    @staticmethod
    def _is_sequence_number(value) -> bool:
        return isinstance(value, Real) and not isinstance(value, bool) and float(value).is_integer()

    def _prepare_data_row(self, row: int) -> None:
        """Extend the form while preserving the preceding row's formatting."""
        if row <= self._sheet.max_row and self._sheet.cell(row=row, column=2).value is not None:
            return
        source_row = max(1, row - 1)
        for column in range(2, max(8, self._sheet.max_column) + 1):
            source = self._sheet.cell(row=source_row, column=column)
            target = self._sheet.cell(row=row, column=column)
            if source.has_style:
                target._style = copy(source._style)
            target.alignment = copy(source.alignment)
        self._sheet.row_dimensions[row].height = self._sheet.row_dimensions[source_row].height
        previous = self._sheet.cell(row=source_row, column=2).value
        sequence = int(previous) + 1 if self._is_sequence_number(previous) else self._appended_count + 1
        self._sheet.cell(row=row, column=2, value=sequence)

    def _find_last_data_row(self) -> int:
        ws = self._sheet
        last_row = 1
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            if any((cell.value is not None for cell in row)):
                last_row = row[0].row
        return last_row

    def append_record(self, url: str, target_brand: str, features: str, detected_at: Optional[str]=None,
                      category: str = '偽ショッピングサイト') -> None:
        ws = self._sheet
        row = self._next_row
        self._prepare_data_row(row)
        date_value = datetime.now().date()
        if detected_at:
            try:
                date_value = datetime.fromisoformat(detected_at.replace('Z', '+00:00')).date()
            except ValueError:
                try:
                    date_value = datetime.strptime(detected_at[:10], '%Y/%m/%d').date()
                except ValueError:
                    pass
        data_map = {'実施年月日': date_value, 'ＳＮＳ別': 'サイト', 'サイト名・ユーザー名': target_brand, 'ＵＲＬ': url, '該当項目': category, '備考': features}
        for col_name, value in data_map.items():
            col_idx = self._column_map.get(col_name)
            if col_idx is None:
                continue
            cell = ws.cell(row=row, column=col_idx, value=value)
            if col_name == '実施年月日':
                cell.number_format = 'yyyy/m/d'
            if col_name == 'ＵＲＬ':
                # Keep suspicious URLs as plain text to prevent accidental access.
                cell.hyperlink = None
        self._next_row += 1
        self._appended_count += 1
        logger.info(f'📝 Excel に追記 (行 {row}): {url[:60]}...')

    def save(self) -> str:
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = self._output_path.with_stem(f'{self._output_path.stem}_{timestamp}')
            self._workbook.save(str(output_path))
            logger.info(f'✅ Excel レポートを保存しました: {output_path} ({self._appended_count} 件追記)')
            return str(output_path)
        except Exception as e:
            logger.error(f'Excel 保存エラー: {e}')
            raise

    @property
    def appended_count(self) -> int:
        return self._appended_count
